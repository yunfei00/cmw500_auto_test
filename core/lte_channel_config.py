from __future__ import annotations

import logging
import math
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from core.channel_config import normalize_band
from core.paths import resource_path, user_data_dir

logger = logging.getLogger(__name__)

LTE_SHEET_NAME = "LTE"

REQUIRED_HEADERS: tuple[str, ...] = (
    "band",
    "固定信道",
    "begin",
    "end",
    "step",
    "线损",
    "bw",
    "转盘bw",
    "转盘",
    "top bw",
    "top",
    "三信道bw",
    "三信道",
)

FIXED_CHANNEL_TEST_ITEMS = frozenset({"普通测试", "固定信道测试"})

DEFAULT_BAND_ROWS: tuple[dict[str, Any], ...] = (
    {
        "band": "B1",
        "固定信道": "是",
        "begin": "",
        "end": "",
        "step": "",
        "线损": 1.5,
        "bw": 20,
        "转盘bw": 20,
        "转盘": "300",
        "top bw": 10,
        "top": "100",
        "三信道bw": 20,
        "三信道": "0,300,599",
    },
    {
        "band": "B3",
        "固定信道": "是",
        "begin": "",
        "end": "",
        "step": "",
        "线损": 2.0,
        "bw": 20,
        "转盘bw": 20,
        "转盘": "1575",
        "top bw": 20,
        "top": "1300",
        "三信道bw": 20,
        "三信道": "1200,1575,1949",
    },
    {
        "band": "B5",
        "固定信道": "是",
        "begin": "",
        "end": "",
        "step": "",
        "线损": 1.8,
        "bw": 10,
        "转盘bw": 10,
        "转盘": "2525",
        "top bw": 10,
        "top": "2450",
        "三信道bw": 10,
        "三信道": "2400,2525,2649",
    },
)


class LTEChannelConfigError(Exception):
    """LTE 信道配置相关错误。"""


@dataclass
class LTEBandChannelConfig:
    band: str
    is_fixed_channel: bool
    begin: int | None
    end: int | None
    step: int | None
    loss_db: float
    bw: float
    turntable_bw: float
    turntable_channels: list[int]
    top_bw: float
    top_channels: list[int]
    three_channel_bw: float
    three_channels: list[int]

    def channels_from_range(self) -> list[int]:
        if self.begin is None or self.end is None or self.step is None:
            return []
        if self.step <= 0:
            return []
        channels: list[int] = []
        current = self.begin
        while current <= self.end:
            channels.append(current)
            current += self.step
        return channels

    def resolved_traverse_channels(self) -> list[int]:
        return self.channels_from_range()


@dataclass
class LteTestChannelSelection:
    bw: float
    channels: list[int]
    loss_db: float


def default_lte_channel_config_path() -> Path:
    return user_data_dir() / "configs" / "lte_channel_config.xlsx"


def parse_int_list(value: Any) -> list[int]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    text = text.replace("，", ",")
    result: list[int] = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            result.append(int(float(part)))
        except (TypeError, ValueError) as exc:
            raise LTEChannelConfigError(f"无法解析整数列表：{value}") from exc
    return result


def parse_optional_int(value: Any) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError) as exc:
        raise LTEChannelConfigError(f"无法解析整数：{value}") from exc


def parse_optional_int_list(value: Any) -> list[int]:
    return parse_int_list(value)


def parse_float_value(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    text = str(value).strip()
    if not text:
        return default
    try:
        parsed = float(text)
    except (TypeError, ValueError) as exc:
        raise LTEChannelConfigError(f"无法解析浮点数：{value}") from exc
    if not math.isfinite(parsed):
        raise LTEChannelConfigError(f"数值必须是有限值：{value}")
    return parsed


def parse_yes_no(value: Any) -> bool:
    if value is None:
        raise LTEChannelConfigError("固定信道列不能为空，应填写“是”或“否”。")
    text = str(value).strip().lower()
    mapping = {
        "是": True,
        "否": False,
        "yes": True,
        "no": False,
        "y": True,
        "n": False,
        "true": True,
        "false": False,
        "1": True,
        "0": False,
    }
    if text in mapping:
        return mapping[text]
    raise LTEChannelConfigError(f"固定信道列无法解析：{value}，请填写“是”或“否”。")


def _normalize_header(value: Any) -> str:
    return str(value).strip()


class LTEChannelConfigManager:
    def __init__(self, path: Path | str | None = None) -> None:
        self.path = Path(path) if path else default_lte_channel_config_path()
        self._bands: dict[str, LTEBandChannelConfig] = {}
        self._load_errors: list[str] = []

    @property
    def load_errors(self) -> list[str]:
        return list(self._load_errors)

    def ensure_default_file(self) -> Path:
        if self.path.exists():
            return self.path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        bundled_path = resource_path("configs/lte_channel_config.xlsx")
        if bundled_path.is_file() and bundled_path.resolve() != self.path.resolve():
            shutil.copy2(bundled_path, self.path)
            logger.info("已复制内置 LTE 信道配置文件：%s", self.path)
        else:
            write_default_lte_channel_excel(self.path)
            logger.info("已自动生成默认 LTE 信道配置文件：%s", self.path)
        return self.path

    def load(self, path: Path | str | None = None) -> None:
        if path is not None:
            self.path = Path(path)
        self._bands.clear()
        self._load_errors.clear()

        if not self.path.exists():
            raise LTEChannelConfigError(f"配置文件不存在：{self.path}")

        from openpyxl import load_workbook

        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            if LTE_SHEET_NAME not in workbook.sheetnames:
                raise LTEChannelConfigError(f"Excel 缺少 Sheet：{LTE_SHEET_NAME}")

            sheet = workbook[LTE_SHEET_NAME]
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                raise LTEChannelConfigError("LTE Sheet 为空")

            header_map = self._build_header_map(header_row)
            missing_headers = [name for name in REQUIRED_HEADERS if name not in header_map]
            if missing_headers:
                message = f"Excel 表头缺失字段：{', '.join(missing_headers)}"
                self._load_errors.append(message)
                raise LTEChannelConfigError(message)

            for row_index, row in enumerate(rows, start=2):
                try:
                    band_config = self._parse_row(row, header_map)
                except LTEChannelConfigError as exc:
                    error = f"第 {row_index} 行解析失败：{exc}"
                    self._load_errors.append(error)
                    logger.warning(error)
                    continue
                if band_config is None:
                    continue
                self._bands[band_config.band] = band_config
        finally:
            workbook.close()

        if not self._bands:
            raise LTEChannelConfigError("未解析到任何有效的 LTE Band 配置")

    def get_all_bands(self) -> list[str]:
        return sorted(self._bands.keys(), key=_band_sort_key)

    def has_config(self) -> bool:
        return bool(self._bands)

    def get_band_config(self, band: str) -> LTEBandChannelConfig:
        band_key = normalize_band(band)
        config = self._bands.get(band_key)
        if config is None:
            raise KeyError(band_key)
        return config

    def get_fixed_channel_selection(self, band: str) -> LteTestChannelSelection | None:
        return None

    def get_band_test_selections(
        self,
        band: str,
        optional_test_items: list[str],
    ) -> list[tuple[str, LteTestChannelSelection]]:
        selections: list[tuple[str, LteTestChannelSelection]] = []
        config = self.get_band_config(band)
        if not config.is_fixed_channel:
            selections.append(("普通测试", self.get_channels_for_test_item(band, "普通测试")))
            return selections
        for test_item in optional_test_items:
            selections.append((test_item, self.get_channels_for_test_item(band, test_item)))
        return selections

    def get_channels_for_test_item(self, band: str, test_item: str) -> LteTestChannelSelection:
        config = self.get_band_config(band)
        item = test_item.strip()
        if not item:
            raise ValueError("当前 Band 未配置该测试项信道。")

        if item in FIXED_CHANNEL_TEST_ITEMS:
            channels = config.resolved_traverse_channels()
            if not channels:
                raise ValueError("当前 Band 未配置该测试项信道。")
            return LteTestChannelSelection(bw=config.bw, channels=channels, loss_db=config.loss_db)

        if item == "转盘测试":
            if not config.turntable_channels:
                raise ValueError("当前 Band 未配置该测试项信道。")
            return LteTestChannelSelection(
                bw=config.turntable_bw,
                channels=list(config.turntable_channels),
                loss_db=config.loss_db,
            )

        if item == "TOP测试":
            if not config.top_channels:
                raise ValueError("当前 Band 未配置该测试项信道。")
            return LteTestChannelSelection(
                bw=config.top_bw,
                channels=list(config.top_channels),
                loss_db=config.loss_db,
            )

        if item == "三信道测试":
            if not config.three_channels:
                raise ValueError("当前 Band 未配置该测试项信道。")
            return LteTestChannelSelection(
                bw=config.three_channel_bw,
                channels=list(config.three_channels),
                loss_db=config.loss_db,
            )

        raise ValueError(f"不支持的测试项：{test_item}")

    def _build_header_map(self, header_row: tuple[Any, ...]) -> dict[str, int]:
        header_map: dict[str, int] = {}
        for index, header in enumerate(header_row):
            if header is None:
                continue
            name = _normalize_header(header)
            if name and name not in header_map:
                header_map[name] = index
        return header_map

    def _value_at(self, row: tuple[Any, ...], index: int) -> Any:
        return row[index] if index < len(row) else None

    def _parse_row(
        self,
        row: tuple[Any, ...],
        header_map: dict[str, int],
    ) -> LTEBandChannelConfig | None:
        band_value = self._value_at(row, header_map["band"])
        if band_value in (None, ""):
            return None

        try:
            band = normalize_band(band_value)
            is_fixed_channel = parse_yes_no(self._value_at(row, header_map["固定信道"]))
            begin = parse_optional_int(self._value_at(row, header_map["begin"]))
            end = parse_optional_int(self._value_at(row, header_map["end"]))
            step = parse_optional_int(self._value_at(row, header_map["step"]))
            loss_db = parse_float_value(self._value_at(row, header_map["线损"]))
            bw = parse_float_value(self._value_at(row, header_map["bw"]))
            turntable_bw = parse_float_value(self._value_at(row, header_map["转盘bw"]))
            turntable_channels = parse_optional_int_list(self._value_at(row, header_map["转盘"]))
            top_bw = parse_float_value(self._value_at(row, header_map["top bw"]))
            top_channels = parse_optional_int_list(self._value_at(row, header_map["top"]))
            three_channel_bw = parse_float_value(self._value_at(row, header_map["三信道bw"]))
            three_channels = parse_int_list(self._value_at(row, header_map["三信道"]))
        except (LTEChannelConfigError, ValueError) as exc:
            raise LTEChannelConfigError(str(exc)) from exc

        band_config = LTEBandChannelConfig(
            band=band,
            is_fixed_channel=is_fixed_channel,
            begin=begin,
            end=end,
            step=step,
            loss_db=loss_db,
            bw=bw,
            turntable_bw=turntable_bw,
            turntable_channels=turntable_channels,
            top_bw=top_bw,
            top_channels=top_channels,
            three_channel_bw=three_channel_bw,
            three_channels=three_channels,
        )
        self._validate_band_config(band_config)
        return band_config

    @staticmethod
    def _validate_band_config(config: LTEBandChannelConfig) -> None:
        if config.loss_db < 0:
            raise LTEChannelConfigError("线损不能为负数")

        range_values = (config.begin, config.end, config.step)
        if any(value is not None for value in range_values):
            if any(value is None for value in range_values):
                raise LTEChannelConfigError("begin、end、step 必须同时填写")
            assert config.begin is not None and config.end is not None and config.step is not None
            if config.begin < 0 or config.end < 0:
                raise LTEChannelConfigError("信道不能为负数")
            if config.end < config.begin:
                raise LTEChannelConfigError("end 不能小于 begin")
            if config.step <= 0:
                raise LTEChannelConfigError("step 必须大于 0")
            if config.bw <= 0:
                raise LTEChannelConfigError("配置遍历信道时 bw 必须大于 0")

        channel_groups = (
            ("转盘", config.turntable_channels, config.turntable_bw),
            ("top", config.top_channels, config.top_bw),
            ("三信道", config.three_channels, config.three_channel_bw),
        )
        for name, channels, bandwidth in channel_groups:
            if any(channel < 0 for channel in channels):
                raise LTEChannelConfigError(f"{name}信道不能为负数")
            if channels and bandwidth <= 0:
                raise LTEChannelConfigError(f"{name}配置有信道时带宽必须大于 0")


def write_default_lte_channel_excel(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = LTE_SHEET_NAME
    sheet.append(list(REQUIRED_HEADERS))

    for row in DEFAULT_BAND_ROWS:
        sheet.append([row.get(header, "") for header in REQUIRED_HEADERS])

    header_fill = PatternFill("solid", fgColor="E5EBF0")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(max_length + 2, 10)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _band_sort_key(band: str) -> tuple[int, str]:
    match = re.search(r"\d+", band)
    return (int(match.group()) if match else 9999, band)
