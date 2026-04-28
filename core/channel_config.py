from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass
class ChannelPoint:
    rat: str
    band: str
    channel_type: str
    channel: int
    frequency_mhz: float | None = None
    extra: dict = field(default_factory=dict)


class ChannelConfigManager:
    SUPPORTED_SHEETS = {"LTE", "WiFi", "WCDMA", "GSM"}

    HEADER_ALIASES = {
        "rat": "rat",
        "制式": "rat",
        "band": "band",
        "频段": "band",
        "channel_type": "channel_type",
        "channeltype": "channel_type",
        "频点类型": "channel_type",
        "信道类型": "channel_type",
        "channel": "channel",
        "信道": "channel",
        "frequency_mhz": "frequency_mhz",
        "frequencymhz": "frequency_mhz",
        "frequency": "frequency_mhz",
        "freq": "frequency_mhz",
        "频率mhz": "frequency_mhz",
        "频率": "frequency_mhz",
    }

    CHANNEL_TYPE_ALIASES = {
        "低": "低频点",
        "低频": "低频点",
        "低频点": "低频点",
        "low": "低频点",
        "中": "中频点",
        "中频": "中频点",
        "中频点": "中频点",
        "middle": "中频点",
        "mid": "中频点",
        "高": "高频点",
        "高频": "高频点",
        "高频点": "高频点",
        "high": "高频点",
        "top": "Top频点",
        "top频点": "Top频点",
        "toppoint": "Top频点",
        "top_point": "Top频点",
    }

    def __init__(self) -> None:
        self.channels: dict[str, dict[str, dict[str, int]]] = {}
        self.points: list[ChannelPoint] = []

    def load_excel(self, file_path: str) -> None:
        from openpyxl import load_workbook

        workbook_path = Path(file_path)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            loaded_channels: dict[str, dict[str, dict[str, int]]] = {}
            loaded_points: list[ChannelPoint] = []

            for sheet_name in workbook.sheetnames:
                rat_name = sheet_name.strip()
                if rat_name.upper() not in self.SUPPORTED_SHEETS:
                    continue
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                header_row = next(rows, None)
                if not header_row:
                    continue

                header_map = self._build_header_map(header_row)
                if not self._has_required_headers(header_map):
                    continue

                for row in rows:
                    point = self._parse_row(row, header_map, rat_name)
                    if not point:
                        continue
                    rat = point.rat.upper()
                    loaded_channels.setdefault(rat, {}).setdefault(point.band, {})[
                        point.channel_type
                    ] = point.channel
                    loaded_points.append(point)

            self.channels = loaded_channels
            self.points = loaded_points
        finally:
            workbook.close()

    def get_channel(self, rat: str, band: str, channel_type: str) -> int | None:
        rat_key = self._normalize_rat(rat)
        band_key = normalize_band(band)
        channel_type_key = normalize_channel_type(channel_type)
        return self.channels.get(rat_key, {}).get(band_key, {}).get(channel_type_key)

    def get_channels(self, rat: str, band: str) -> dict[str, int]:
        rat_key = self._normalize_rat(rat)
        band_key = normalize_band(band)
        return dict(self.channels.get(rat_key, {}).get(band_key, {}))

    def get_supported_bands(self, rat: str) -> list[str]:
        rat_key = self._normalize_rat(rat)
        return sorted(self.channels.get(rat_key, {}).keys(), key=_band_sort_key)

    def has_config(self) -> bool:
        return any(self.channels.values())

    def _build_header_map(self, header_row: tuple[Any, ...]) -> dict[str, int]:
        header_map: dict[str, int] = {}
        for index, header in enumerate(header_row):
            if header is None:
                continue
            normalized_header = self._normalize_header(str(header))
            canonical = self.HEADER_ALIASES.get(normalized_header)
            if canonical and canonical not in header_map:
                header_map[canonical] = index
        return header_map

    def _has_required_headers(self, header_map: dict[str, int]) -> bool:
        return {"band", "channel_type", "channel"}.issubset(header_map)

    def _parse_row(
        self,
        row: tuple[Any, ...],
        header_map: dict[str, int],
        default_rat: str,
    ) -> ChannelPoint | None:
        band_value = self._value_at(row, header_map["band"])
        channel_type_value = self._value_at(row, header_map["channel_type"])
        channel_value = self._value_at(row, header_map["channel"])
        if band_value in (None, "") or channel_type_value in (None, "") or channel_value in (None, ""):
            return None

        rat_value = self._value_at(row, header_map["rat"]) if "rat" in header_map else default_rat
        frequency_value = (
            self._value_at(row, header_map["frequency_mhz"])
            if "frequency_mhz" in header_map
            else None
        )

        try:
            band = normalize_band(band_value)
            channel_type = normalize_channel_type(channel_type_value)
            channel = int(float(channel_value))
            frequency_mhz = (
                None if frequency_value in (None, "") else float(frequency_value)
            )
        except (TypeError, ValueError):
            return None

        return ChannelPoint(
            rat=self._normalize_rat(str(rat_value or default_rat)),
            band=band,
            channel_type=channel_type,
            channel=channel,
            frequency_mhz=frequency_mhz,
        )

    def _normalize_header(self, value: str) -> str:
        return value.strip().lower().replace(" ", "").replace("-", "_")

    def _normalize_rat(self, value: str) -> str:
        return value.strip().upper()

    def _value_at(self, row: tuple[Any, ...], index: int) -> Any:
        return row[index] if index < len(row) else None


def normalize_band(value: object) -> str:
    text = str(value).strip().upper()
    if not text:
        raise ValueError("empty band")
    if re.fullmatch(r"\d+(\.0+)?", text):
        return f"B{int(float(text))}"
    match = re.search(r"B\s*(\d+)", text)
    if match:
        return f"B{int(match.group(1))}"
    match = re.search(r"(\d+)", text)
    if match:
        return f"B{int(match.group(1))}"
    raise ValueError(f"invalid band: {value}")


def normalize_channel_type(value: object) -> str:
    text = str(value).strip()
    if not text:
        raise ValueError("empty channel type")
    normalized = text.lower().replace(" ", "").replace("-", "_")
    if normalized in ChannelConfigManager.CHANNEL_TYPE_ALIASES:
        return ChannelConfigManager.CHANNEL_TYPE_ALIASES[normalized]
    raise ValueError(f"invalid channel type: {value}")


def _band_sort_key(band: str) -> tuple[int, str]:
    match = re.search(r"\d+", band)
    return (int(match.group()) if match else 9999, band)
