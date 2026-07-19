from __future__ import annotations

import json
import os
from pathlib import Path
import tempfile

from core.models import TestResult
from core.result_summary import SummaryResult


RAW_HEADERS = [
    "Run ID",
    "序号",
    "数据来源",
    "制式",
    "Band",
    "信道",
    "频点类型",
    "测试模式",
    "带宽(MHz)",
    "DUT目标电平(dBm)",
    "仪表下发电平(dBm)",
    "全局线损(dB)",
    "信道线损(dB)",
    "总线损(dB)",
    "指标类型",
    "指标值",
    "包数",
    "BLER门限(%)",
    "灵敏度规格上限(dBm)",
    "尝试次数",
    "扫描阶段",
    "结果",
    "状态",
    "错误信息",
    "时间",
]

SUMMARY_HEADERS = [
    "Run ID",
    "数据来源",
    "制式",
    "Band",
    "信道",
    "频点类型",
    "测试模式",
    "灵敏度(dBm)",
    "规格上限(dBm)",
    "PASS数量",
    "FAIL数量",
    "总数",
    "结果",
    "备注",
]


def export_results_to_excel(
    raw_results: list[TestResult],
    summary_results: list[SummaryResult],
    file_path: str,
    run_metadata: dict | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    raw_sheet = workbook.active
    raw_sheet.title = "RawResults"
    summary_sheet = workbook.create_sheet("Summary")
    metadata_sheet = workbook.create_sheet("RunMetadata")
    trace_sheet = workbook.create_sheet("SCPITrace")

    fills = {
        "PASS": PatternFill("solid", fgColor="EAF7EA"),
        "FAIL": PatternFill("solid", fgColor="FDECEC"),
        "ERROR": PatternFill("solid", fgColor="FFF0C2"),
        "FAILED": PatternFill("solid", fgColor="FFF0C2"),
        "FAILED_UNSAFE": PatternFill("solid", fgColor="F8C4C4"),
        "STOPPED": PatternFill("solid", fgColor="FFF0C2"),
    }
    header_fill = PatternFill("solid", fgColor="E5EBF0")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    metadata = dict(run_metadata or {})
    if any(
        str(getattr(result, "data_source", "")).upper() == "SIMULATION"
        for result in raw_results
    ):
        metadata["contains_simulation_data"] = True
    warnings = _run_warnings(metadata)
    banner = " | ".join(warnings)

    _write_sheet(
        raw_sheet,
        RAW_HEADERS,
        [_raw_result_row(result) for result in raw_results],
        fills,
        header_fill,
        header_font,
        center,
        result_column_name="结果",
        banner=banner,
    )
    _write_sheet(
        summary_sheet,
        SUMMARY_HEADERS,
        [_summary_result_row(result) for result in summary_results],
        fills,
        header_fill,
        header_font,
        center,
        result_column_name="结果",
        banner=banner,
    )

    _write_metadata_sheet(metadata_sheet, metadata, warnings, header_fill, header_font)
    _write_trace_sheet(trace_sheet, metadata.get("command_trace", []), header_fill, header_font)

    output_path = Path(file_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{output_path.stem}_",
            suffix=".xlsx",
            dir=output_path.parent,
            delete=False,
        ) as temp_file:
            temp_path = Path(temp_file.name)
        workbook.save(temp_path)
        os.replace(temp_path, output_path)
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink()


def _write_sheet(
    sheet,
    headers: list[str],
    rows: list[list[object]],
    fills: dict[str, object],
    header_fill,
    header_font,
    center,
    result_column_name: str,
    banner: str = "",
) -> None:
    from openpyxl.styles import Font, PatternFill

    header_row = 1
    if banner:
        sheet.append([banner])
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        banner_cell = sheet.cell(1, 1)
        banner_cell.font = Font(bold=True, color="FFFFFF")
        banner_cell.fill = PatternFill("solid", fgColor="9B1C1C")
        banner_cell.alignment = center
        header_row = 2
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_safe_value(value) for value in row])

    result_column_index = headers.index(result_column_name) + 1
    for cell in sheet[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        result_value = str(row[result_column_index - 1].value or "").upper()
        row_fill = fills.get(result_value)
        for cell in row:
            cell.alignment = center
            if row_fill:
                cell.fill = row_fill

    sheet.freeze_panes = f"A{header_row + 1}"
    _auto_fit_columns(sheet)


def _auto_fit_columns(sheet) -> None:
    from openpyxl.utils import get_column_letter

    for column_index, column_cells in enumerate(sheet.iter_cols(), start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 32)


def _raw_result_row(result: TestResult) -> list[object]:
    return [
        getattr(result, "run_id", ""),
        result.index,
        getattr(result, "data_source", ""),
        result.mode,
        result.band,
        result.channel,
        result.channel_type,
        result.test_mode,
        getattr(result, "bw", None),
        result.rx_level,
        getattr(result, "instrument_level", None),
        getattr(result, "global_cable_loss", None),
        getattr(result, "channel_loss", None),
        getattr(result, "total_loss", None),
        result.metric_type,
        result.metric_value,
        getattr(result, "packet_count", None),
        getattr(result, "bler_threshold", None),
        getattr(result, "sensitivity_upper", None),
        getattr(result, "attempt", 1),
        getattr(result, "scan_phase", ""),
        result.result,
        result.status,
        getattr(result, "error_message", ""),
        result.timestamp,
    ]


def _summary_result_row(result: SummaryResult) -> list[object]:
    return [
        getattr(result, "run_id", ""),
        getattr(result, "data_source", ""),
        result.mode,
        result.band,
        result.channel,
        result.channel_type,
        result.test_mode,
        "-" if result.sensitivity is None else result.sensitivity,
        getattr(result, "sensitivity_upper", None),
        result.pass_count,
        result.fail_count,
        result.total_count,
        result.result,
        result.remark,
    ]


def _write_metadata_sheet(
    sheet,
    metadata: dict,
    warnings: list[str],
    header_fill,
    header_font,
) -> None:
    from openpyxl.styles import Font, PatternFill

    sheet.append(["Field", "Value"])
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for warning in warnings:
        sheet.append(["WARNING", warning])
        for cell in sheet[sheet.max_row]:
            cell.fill = PatternFill("solid", fgColor="9B1C1C")
            cell.font = Font(bold=True, color="FFFFFF")
    for key, value in metadata.items():
        if key == "command_trace":
            continue
        if isinstance(value, (dict, list)):
            rendered = json.dumps(value, ensure_ascii=False, sort_keys=True)
        else:
            rendered = value
        sheet.append([_excel_safe_text(str(key)), _excel_safe_text(str(rendered))])
    sheet.freeze_panes = "A2"
    _auto_fit_columns(sheet)


def _write_trace_sheet(sheet, trace: object, header_fill, header_font) -> None:
    headers = ["timestamp", "stage", "operation", "command", "response", "success", "error"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    if isinstance(trace, list):
        for entry in trace:
            if not isinstance(entry, dict):
                continue
            sheet.append([_excel_safe_text(str(entry.get(header, ""))) for header in headers])
    sheet.freeze_panes = "A2"
    _auto_fit_columns(sheet)


def _excel_safe_text(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _excel_safe_value(value: object) -> object:
    if isinstance(value, str):
        return _excel_safe_text(value)
    return value


def _run_warnings(metadata: dict) -> list[str]:
    warnings: list[str] = []
    if (
        str(metadata.get("data_source", "")).upper() == "SIMULATION"
        or bool(metadata.get("contains_simulation_data", False))
    ):
        warnings.append("SIMULATED DATA - 非实测结果，不得作为正式报告")

    status = str(metadata.get("status", "")).strip().upper()
    if status and status != "COMPLETED":
        if status == "FAILED_UNSAFE":
            warnings.append(
                "UNSAFE / INCOMPLETE RUN - 仪表安全清理未确认，结果无效，必须人工确认 RF 状态"
            )
        else:
            warnings.append(
                f"INCOMPLETE RUN ({status}) - 测试未正常完成，结果不得作为正式 PASS 结论"
            )
    return warnings
