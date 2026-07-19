from __future__ import annotations

import hashlib
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app_info import APP_VERSION, load_app_version, load_build_info
from core.lte_channel_config import REQUIRED_HEADERS
from scripts import build_windows, package_release


PROJECT_ROOT = Path(__file__).resolve().parents[1]


@pytest.mark.parametrize("version", ["dev", "v0.1.0", "v10.20.30"])
def test_validate_version_accepts_supported_versions(version: str) -> None:
    assert package_release.validate_version(version) == version


@pytest.mark.parametrize(
    "version",
    ["", "1.2.3", "v1.2", "v1.2.3.4", "v01.2.3", "v1.02.3", "feature/test"],
)
def test_validate_version_rejects_unsafe_or_non_release_versions(version: str) -> None:
    with pytest.raises(ValueError, match="vX.Y.Z"):
        package_release.validate_version(version)


def test_resolve_version_uses_only_explicit_or_tag_versions(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("GITHUB_REF_TYPE", "branch")
    monkeypatch.setenv("GITHUB_REF_NAME", "feature/release-test")
    assert package_release.resolve_version(None) == "dev"

    monkeypatch.setenv("GITHUB_REF_TYPE", "tag")
    monkeypatch.setenv("GITHUB_REF_NAME", "v2.3.4")
    assert package_release.resolve_version(None) == "v2.3.4"
    assert package_release.resolve_version("v9.8.7") == "v9.8.7"


def test_build_runtime_version_is_bound_to_release_tag(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("GITHUB_REF_TYPE", "tag")
    monkeypatch.setenv("GITHUB_REF_NAME", "v2.3.4")
    monkeypatch.delenv(build_windows.BUILD_VERSION_ENV, raising=False)
    assert build_windows.resolve_runtime_version() == "v2.3.4"

    monkeypatch.setenv(build_windows.BUILD_VERSION_ENV, "v2.3.5")
    with pytest.raises(ValueError, match="does not match release tag"):
        build_windows.resolve_runtime_version()


def test_windows_version_resource_contains_runtime_identity(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    version_info_path = tmp_path / "version_info.txt"
    monkeypatch.setattr(build_windows, "VERSION_INFO_FILE", version_info_path)
    monkeypatch.setattr(build_windows, "resolve_runtime_version", lambda: "v2.3.4")
    monkeypatch.setattr(
        build_windows,
        "resolve_source_identity",
        lambda: ("a" * 40, False),
    )

    build_windows.write_windows_version_info()

    content = version_info_path.read_text(encoding="utf-8")
    assert "filevers=(2, 3, 4, 0)" in content
    assert "ProductVersion', 'v2.3.4'" in content
    assert "Source " + ("a" * 40) in content


def test_sample_channel_config_uses_current_schema() -> None:
    workbook = load_workbook(
        PROJECT_ROOT / "sample_channel_config.xlsx",
        read_only=True,
        data_only=True,
    )
    try:
        headers = tuple(next(workbook["LTE"].iter_rows(values_only=True)))
    finally:
        workbook.close()

    assert headers == REQUIRED_HEADERS
    assert len(headers) == 13


def test_app_version_comes_from_required_version_resource(tmp_path: Path) -> None:
    assert APP_VERSION == (PROJECT_ROOT / "VERSION").read_text(encoding="ascii").strip()

    release_version = tmp_path / "VERSION"
    release_version.write_text("v2.3.4\n", encoding="ascii")
    assert load_app_version(release_version) == "v2.3.4"

    release_version.write_text("not-a-release\n", encoding="ascii")
    with pytest.raises(RuntimeError, match="Invalid VERSION"):
        load_app_version(release_version)


def test_build_info_is_required_and_bound_to_app_version(tmp_path: Path) -> None:
    build_info_path = tmp_path / "BUILD_INFO.json"
    build_info_path.write_text(
        '{"version": "'
        + APP_VERSION
        + '", "commit": "'
        + ("b" * 40)
        + '", "built_at": "2026-07-19T10:00:00+00:00", "dirty": false}',
        encoding="utf-8",
    )

    value = load_build_info(build_info_path)
    assert value["commit"] == "b" * 40

    build_info_path.write_text(
        '{"version": "v9.9.9", "commit": "'
        + ("b" * 40)
        + '", "built_at": "2026-07-19T10:00:00+00:00", "dirty": false}',
        encoding="utf-8",
    )
    with pytest.raises(RuntimeError, match="does not match VERSION"):
        load_build_info(build_info_path)


def test_verify_runtime_resources_requires_every_critical_path(tmp_path: Path) -> None:
    for relative_path in build_windows.REQUIRED_RUNTIME_PATHS:
        target = tmp_path / relative_path
        if relative_path == Path("_internal/pyvisa"):
            target.mkdir(parents=True)
        else:
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(b"test")

    build_windows.verify_runtime_resources(tmp_path)

    missing_path = tmp_path / "configs/lte_channel_config.xlsx"
    missing_path.unlink()
    with pytest.raises(FileNotFoundError, match="configs.lte_channel_config.xlsx"):
        build_windows.verify_runtime_resources(tmp_path)


def test_verify_bundled_modules_requires_pyvisa(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from PyInstaller.archive import readers

    archive_path = tmp_path / "PYZ-00.pyz"
    archive_path.write_bytes(b"test")

    class FakeReader:
        def __init__(self, _path: str) -> None:
            self.toc = {
                "pyvisa": object(),
                "pyvisa_py": object(),
                "psutil": object(),
                "other_module": object(),
            }

    monkeypatch.setattr(readers, "ZlibArchiveReader", FakeReader)
    build_windows.verify_bundled_modules(archive_path)

    class MissingReader:
        def __init__(self, _path: str) -> None:
            self.toc = {"other_module": object()}

    monkeypatch.setattr(readers, "ZlibArchiveReader", MissingReader)
    with pytest.raises(RuntimeError, match="pyvisa"):
        build_windows.verify_bundled_modules(archive_path)


def _create_minimal_release_tree(
    app_dist_dir: Path,
    runtime_version: str,
) -> None:
    for archive_name in package_release.REQUIRED_ARCHIVE_FILES:
        relative_path = Path(archive_name).relative_to(package_release.APP_NAME)
        target = app_dist_dir / relative_path
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(f"content for {archive_name}".encode("utf-8"))

    (app_dist_dir / "VERSION").write_text(
        f"{runtime_version}\n",
        encoding="ascii",
        newline="\n",
    )
    (app_dist_dir / "BUILD_INFO.json").write_text(
        '{"version": "'
        + runtime_version
        + '", "commit": "'
        + ("a" * 40)
        + '", "built_at": "2026-07-19T10:00:00+00:00", "dirty": false}\n',
        encoding="utf-8",
    )


def test_create_zip_verifies_resources_and_writes_sha256_manifest(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    app_dist_dir = tmp_path / "dist" / package_release.APP_NAME
    release_dir = tmp_path / "release"
    _create_minimal_release_tree(app_dist_dir, "v1.2.3")
    monkeypatch.setattr(package_release, "APP_DIST_DIR", app_dist_dir)
    monkeypatch.setattr(package_release, "RELEASE_DIR", release_dir)
    verified_signatures: list[Path] = []
    monkeypatch.setattr(
        package_release,
        "verify_authenticode_signature",
        verified_signatures.append,
    )

    zip_path = package_release.create_zip("v1.2.3")
    manifest_path = package_release.create_sha256_manifest(zip_path)

    with zipfile.ZipFile(zip_path, "r") as archive:
        members = {name.replace("\\", "/") for name in archive.namelist()}
    assert set(package_release.REQUIRED_ARCHIVE_FILES).issubset(members)
    assert verified_signatures == [app_dist_dir / f"{package_release.APP_NAME}.exe"]

    expected_hash = hashlib.sha256(zip_path.read_bytes()).hexdigest()
    assert manifest_path.read_text(encoding="ascii") == f"{expected_hash}  {zip_path.name}\n"


def test_create_zip_fails_when_a_critical_resource_is_missing(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    app_dist_dir = tmp_path / "dist" / package_release.APP_NAME
    _create_minimal_release_tree(app_dist_dir, "0.2.0-dev")
    (app_dist_dir / "sample_channel_config.xlsx").unlink()
    monkeypatch.setattr(package_release, "APP_DIST_DIR", app_dist_dir)
    monkeypatch.setattr(package_release, "RELEASE_DIR", tmp_path / "release")

    with pytest.raises(FileNotFoundError, match="sample_channel_config.xlsx"):
        package_release.create_zip("dev")


def test_release_package_fails_when_runtime_version_does_not_match_tag(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    app_dist_dir = tmp_path / "dist" / package_release.APP_NAME
    _create_minimal_release_tree(app_dist_dir, "v1.2.3")
    monkeypatch.setattr(package_release, "APP_DIST_DIR", app_dist_dir)
    monkeypatch.setattr(package_release, "RELEASE_DIR", tmp_path / "release")

    with pytest.raises(RuntimeError, match="does not match"):
        package_release.create_zip("v1.2.4")


def test_release_package_rejects_dirty_build_identity(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    app_dist_dir = tmp_path / "dist" / package_release.APP_NAME
    _create_minimal_release_tree(app_dist_dir, "v1.2.3")
    build_info_path = app_dist_dir / "BUILD_INFO.json"
    build_info = build_info_path.read_text(encoding="utf-8").replace(
        '"dirty": false', '"dirty": true'
    )
    build_info_path.write_text(build_info, encoding="utf-8")
    monkeypatch.setattr(package_release, "APP_DIST_DIR", app_dist_dir)
    monkeypatch.setattr(package_release, "RELEASE_DIR", tmp_path / "release")

    with pytest.raises(RuntimeError, match="dirty source tree"):
        package_release.create_zip("v1.2.3")


def test_signing_can_be_made_mandatory_without_storing_a_certificate(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.delenv(build_windows.SIGNTOOL_PATH_ENV, raising=False)
    monkeypatch.delenv(build_windows.SIGN_CERT_SHA1_ENV, raising=False)
    monkeypatch.setenv(build_windows.REQUIRE_SIGNING_ENV, "1")

    with pytest.raises(RuntimeError, match="Signing is required"):
        build_windows.sign_windows_executable()

    monkeypatch.delenv(build_windows.REQUIRE_SIGNING_ENV, raising=False)
    monkeypatch.setattr(build_windows, "resolve_runtime_version", lambda: "v1.2.3")
    with pytest.raises(RuntimeError, match="Signing is required"):
        build_windows.sign_windows_executable()

    monkeypatch.setenv(build_windows.SIGNTOOL_PATH_ENV, "signtool")
    monkeypatch.setenv(build_windows.SIGN_CERT_SHA1_ENV, "invalid")
    with pytest.raises(ValueError, match="40 hexadecimal"):
        build_windows.sign_windows_executable()


@pytest.mark.parametrize("filename", ["requirements.txt", "requirements-build.txt"])
def test_direct_dependencies_are_exactly_pinned(filename: str) -> None:
    lines = [
        line.strip()
        for line in (PROJECT_ROOT / filename).read_text(encoding="utf-8").splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    ]
    assert lines
    assert all(line.count("==") == 1 for line in lines)
