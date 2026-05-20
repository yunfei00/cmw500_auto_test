from __future__ import annotations

from pathlib import Path

import pytest

from core.lte_channel_config import (
    LTEChannelConfigError,
    LTEChannelConfigManager,
    parse_int_list,
    write_default_lte_channel_excel,
)
from core.models import LteTestConfig
from core.test_plan import generate_lte_test_plan


def _sample_config(**kwargs: object) -> LteTestConfig:
    defaults = {
        "cable_loss": 0.0,
        "sensitivity_upper": -70.0,
        "start_level": -70.0,
        "stop_level": -72.0,
        "packet_count": 1000,
        "max_step": 4.0,
        "min_step": 1.0,
        "bler_threshold": 10.0,
        "settle_time": 2,
        "retry_count": 1,
        "selected_bands": ["B3"],
        "selected_channel_types": [],
        "custom_channels": [],
        "lte_test_items": [],
        "test_mode": "单主",
    }
    defaults.update(kwargs)
    return LteTestConfig(**defaults)


def test_parse_int_list_supports_chinese_comma() -> None:
    assert parse_int_list("1200，1575，1949") == [1200, 1575, 1949]


def test_ensure_default_file_and_load(tmp_path: Path) -> None:
    config_path = tmp_path / "lte_channel_config.xlsx"
    manager = LTEChannelConfigManager(config_path)
    manager.ensure_default_file()
    assert config_path.exists()

    manager.load()
    bands = manager.get_all_bands()
    assert bands == ["B1", "B3", "B5"]


def test_get_fixed_and_optional_channels(tmp_path: Path) -> None:
    config_path = tmp_path / "lte_channel_config.xlsx"
    write_default_lte_channel_excel(config_path)
    manager = LTEChannelConfigManager(config_path)
    manager.load()

    fixed = manager.get_fixed_channel_selection("B3")
    assert fixed is not None
    assert fixed.bw == 20
    assert fixed.channels == [1200, 1575, 1949]

    turntable = manager.get_channels_for_test_item("B3", "转盘测试")
    assert turntable.bw == 20
    assert turntable.channels == [1575]

    top = manager.get_channels_for_test_item("B3", "TOP测试")
    assert top.bw == 20
    assert top.channels == [1300]

    three = manager.get_channels_for_test_item("B3", "三信道测试")
    assert three.bw == 20
    assert three.channels == [1200, 1575, 1949]


def test_get_band_test_selections_includes_fixed_by_default(tmp_path: Path) -> None:
    config_path = tmp_path / "lte_channel_config.xlsx"
    write_default_lte_channel_excel(config_path)
    manager = LTEChannelConfigManager(config_path)
    manager.load()

    selections = manager.get_band_test_selections("B3", [])
    assert len(selections) == 1
    assert selections[0][0] == "固定信道"
    assert selections[0][1].channels == [1200, 1575, 1949]

    multi = manager.get_band_test_selections("B3", ["转盘测试", "TOP测试"])
    assert [name for name, _ in multi] == ["固定信道", "转盘测试", "TOP测试"]


def test_missing_band_raises_key_error(tmp_path: Path) -> None:
    config_path = tmp_path / "lte_channel_config.xlsx"
    write_default_lte_channel_excel(config_path)
    manager = LTEChannelConfigManager(config_path)
    manager.load()

    with pytest.raises(KeyError):
        manager.get_band_config("B99")


def test_empty_test_item_channels_raises_value_error(tmp_path: Path) -> None:
    from openpyxl import Workbook

    config_path = tmp_path / "empty_top.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LTE"
    headers = [
        "band",
        "固定信道",
        "begin",
        "end",
        "step",
        "线损",
        "bw",
        "转盘bw",
        "转盘",
        "top bw",
        "top",
        "三信道bw",
        "三信道",
    ]
    sheet.append(headers)
    sheet.append(["B9", "", "", "", "", 1.0, 20, 20, "", 10, "", 20, ""])
    workbook.save(config_path)

    manager = LTEChannelConfigManager(config_path)
    manager.load()

    with pytest.raises(ValueError, match="当前 Band 未配置该测试项信道"):
        manager.get_channels_for_test_item("B9", "TOP测试")


def test_generate_channels_from_begin_end_step(tmp_path: Path) -> None:
    from openpyxl import Workbook

    config_path = tmp_path / "range.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LTE"
    headers = [
        "band",
        "固定信道",
        "begin",
        "end",
        "step",
        "线损",
        "bw",
        "转盘bw",
        "转盘",
        "top bw",
        "top",
        "三信道bw",
        "三信道",
    ]
    sheet.append(headers)
    sheet.append(["B7", "", 2750, 2760, 5, 1.0, 20, 20, 3100, 10, 3449, 20, ""])
    workbook.save(config_path)

    manager = LTEChannelConfigManager(config_path)
    manager.load()
    selection = manager.get_fixed_channel_selection("B7")
    assert selection is not None
    assert selection.channels == [2750, 2755, 2760]


def test_missing_headers_raises(tmp_path: Path) -> None:
    from openpyxl import Workbook

    config_path = tmp_path / "bad.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LTE"
    sheet.append(["band", "固定信道"])
    sheet.append(["B1", "0,300,599"])
    workbook.save(config_path)

    manager = LTEChannelConfigManager(config_path)
    with pytest.raises(LTEChannelConfigError, match="表头缺失字段"):
        manager.load()


def test_generate_lte_test_plan_from_excel(tmp_path: Path) -> None:
    config_path = tmp_path / "lte_channel_config.xlsx"
    write_default_lte_channel_excel(config_path)
    manager = LTEChannelConfigManager(config_path)
    manager.load()

    config = _sample_config(
        selected_bands=["B3"],
        lte_test_items=[],
        start_level=-70.0,
        stop_level=-70.0,
    )
    items = generate_lte_test_plan(config, manager)
    assert len(items) == 3
    assert sorted(item.channel for item in items) == [1200, 1575, 1949]
    assert all(item.bw == 20 for item in items)

    turntable_config = _sample_config(
        selected_bands=["B3"],
        lte_test_items=["转盘测试"],
        start_level=-70.0,
        stop_level=-70.0,
    )
    turntable_items = generate_lte_test_plan(turntable_config, manager)
    assert len(turntable_items) == 4
    assert turntable_items[-1].channel == 1575
    assert turntable_items[-1].channel_type == "转盘测试"
