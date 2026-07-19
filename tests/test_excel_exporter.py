from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.models import TestResult as Result
from core.result_summary import SummaryResult
from reports.excel_exporter import RAW_HEADERS, SUMMARY_HEADERS, export_results_to_excel


def test_simulation_export_is_traceable_and_visibly_marked(tmp_path: Path) -> None:
    raw_result = Result(
        index=1,
        mode="LTE",
        band="B3",
        channel=1575,
        channel_type="转盘测试",
        test_mode="单主",
        rx_level=-86.0,
        metric_type="BLER",
        metric_value=3.2,
        result="PASS",
        status="COMPLETED",
        run_id="run-sim-001",
        data_source="SIMULATION",
        bw=20.0,
        global_cable_loss=1.0,
        channel_loss=2.0,
        total_loss=3.0,
        instrument_level=-83.0,
        packet_count=1000,
        bler_threshold=10.0,
        sensitivity_upper=-85.0,
        attempt=1,
        scan_phase="FINE",
        error_message="=WEBSERVICE(\"https://invalid.example\")",
    )
    summary_result = SummaryResult(
        mode="LTE",
        band="B3",
        channel=1575,
        channel_type="转盘测试",
        test_mode="单主",
        sensitivity=-86.0,
        pass_count=1,
        fail_count=0,
        total_count=1,
        result="PASS",
        remark="meets specification",
        run_id="run-sim-001",
        data_source="SIMULATION",
        bw=20.0,
        sensitivity_upper=-85.0,
    )
    output_path = tmp_path / "nested" / "result.xlsx"
    metadata = {
        "run_id": "run-sim-001",
        "data_source": "SIMULATION",
        "instrument_idn": "=unsafe-formula",
        "command_trace": [
            {
                "timestamp": "2026-07-19T10:00:00+08:00",
                "stage": "IDN",
                "operation": "query",
                "command": "=unsafe-command",
                "response": "Fake CMW500 Simulator",
                "success": True,
                "error": "",
            }
        ],
    }

    export_results_to_excel(
        [raw_result],
        [summary_result],
        str(output_path),
        run_metadata=metadata,
    )

    workbook = load_workbook(output_path, data_only=False)
    try:
        assert workbook.sheetnames == ["RawResults", "Summary", "RunMetadata", "SCPITrace"]
        assert str(workbook["RawResults"]["A1"].value).startswith("SIMULATED DATA")
        assert [cell.value for cell in workbook["RawResults"][2]] == RAW_HEADERS
        assert workbook["RawResults"]["A3"].value == "run-sim-001"
        assert workbook["RawResults"]["C3"].value == "SIMULATION"
        assert workbook["RawResults"]["X3"].value.startswith("'=WEBSERVICE")
        assert [cell.value for cell in workbook["Summary"][2]] == SUMMARY_HEADERS
        assert workbook["Summary"]["A3"].value == "run-sim-001"
        metadata_values = {
            row[0].value: row[1].value
            for row in workbook["RunMetadata"].iter_rows(min_row=2)
        }
        assert metadata_values["WARNING"].startswith("SIMULATED DATA")
        assert metadata_values["instrument_idn"] == "'=unsafe-formula"
        assert workbook["SCPITrace"]["D2"].value == "'=unsafe-command"
    finally:
        workbook.close()

    assert not list(output_path.parent.glob(f".{output_path.stem}_*.xlsx"))

    mismatch_path = tmp_path / "mismatched_metadata.xlsx"
    export_results_to_excel(
        [raw_result],
        [summary_result],
        str(mismatch_path),
        run_metadata={"data_source": "REAL", "status": "COMPLETED"},
    )
    mismatch_workbook = load_workbook(mismatch_path, read_only=True, data_only=True)
    try:
        assert str(mismatch_workbook["RawResults"]["A1"].value).startswith("SIMULATED DATA")
    finally:
        mismatch_workbook.close()


def test_empty_failed_run_can_still_be_exported(tmp_path: Path) -> None:
    output_path = tmp_path / "failed.xlsx"

    export_results_to_excel(
        [],
        [],
        str(output_path),
        run_metadata={"run_id": "failed-run", "status": "FAILED", "data_source": "REAL"},
    )

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        assert str(workbook["RawResults"]["A1"].value).startswith("INCOMPLETE RUN (FAILED)")
        assert [cell.value for cell in workbook["RawResults"][2]] == RAW_HEADERS
        assert [cell.value for cell in workbook["Summary"][2]] == SUMMARY_HEADERS
        assert workbook["RunMetadata"]["A2"].value == "WARNING"
        assert "INCOMPLETE RUN" in workbook["RunMetadata"]["B2"].value
    finally:
        workbook.close()


def test_failed_unsafe_run_has_strong_safety_warning(tmp_path: Path) -> None:
    output_path = tmp_path / "unsafe.xlsx"

    export_results_to_excel(
        [],
        [],
        str(output_path),
        run_metadata={
            "run_id": "unsafe-run",
            "status": "FAILED_UNSAFE",
            "data_source": "REAL",
        },
    )

    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        assert str(workbook["RawResults"]["A1"].value).startswith("UNSAFE / INCOMPLETE RUN")
        assert "人工确认 RF 状态" in workbook["RunMetadata"]["B2"].value
    finally:
        workbook.close()
