from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.test_plan import DEFAULT_LTE_CHANNELS


SAMPLE_FREQUENCIES = {
    "B1": {"低频点": 2110.0, "中频点": 2140.0, "高频点": 2170.0, "Top频点": 2170.0},
    "B3": {"低频点": 1805.0, "中频点": 1842.5, "高频点": 1880.0, "Top频点": 1880.0},
    "B5": {"低频点": 869.0, "中频点": 881.5, "高频点": 894.0, "Top频点": 894.0},
    "B7": {"低频点": 2620.0, "中频点": 2655.0, "高频点": 2690.0, "Top频点": 2690.0},
    "B8": {"低频点": 925.0, "中频点": 942.5, "高频点": 960.0, "Top频点": 960.0},
    "B20": {"低频点": 791.0, "中频点": 806.0, "高频点": 821.0, "Top频点": 821.0},
    "B28": {"低频点": 758.0, "中频点": 780.5, "高频点": 803.0, "Top频点": 803.0},
    "B38": {"低频点": 2570.0, "中频点": 2595.0, "高频点": 2620.0, "Top频点": 2620.0},
    "B40": {"低频点": 2300.0, "中频点": 2350.0, "高频点": 2400.0, "Top频点": 2400.0},
    "B41": {"低频点": 2496.0, "中频点": 2593.0, "高频点": 2690.0, "Top频点": 2690.0},
}


def main() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LTE"

    headers = ["rat", "band", "channel_type", "channel", "frequency_mhz"]
    sheet.append(headers)

    for band in ["B1", "B3", "B5", "B7", "B8", "B20", "B28", "B38", "B40", "B41"]:
        channel_map = DEFAULT_LTE_CHANNELS[band]
        frequency_map = SAMPLE_FREQUENCIES[band]
        for channel_type in ["低频点", "中频点", "高频点", "Top频点"]:
            sheet.append(
                [
                    "LTE",
                    band,
                    channel_type,
                    channel_map[channel_type],
                    frequency_map[channel_type],
                ]
            )

    header_fill = PatternFill("solid", fgColor="E5EBF0")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(max_length + 2, 12)

    output_path = PROJECT_ROOT / "sample_channel_config.xlsx"
    workbook.save(output_path)
    print(f"Generated: {output_path}")


if __name__ == "__main__":
    main()
